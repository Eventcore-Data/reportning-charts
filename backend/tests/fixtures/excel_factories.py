"""
openpyxl-based helpers that return .xlsx bytes for use in integration tests.
Each function returns a bytes object suitable for use as a file upload payload.
"""

import io
from typing import Optional, Sequence, Tuple

import openpyxl


def make_simple_xlsx(
    headers: Sequence[str] = ("Month", "Revenue"),
    rows: Sequence[Tuple] = (("Jan", 1000), ("Feb", 1500), ("Mar", 1200)),
    title: Optional[str] = None,
) -> bytes:
    """Return bytes of a minimal .xlsx with a single data section."""
    wb = openpyxl.Workbook()
    ws = wb.active
    if title:
        ws.cell(row=1, column=1, value=title)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col_idx, value=header)
        for row_offset, row in enumerate(rows, start=3):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_offset, column=col_idx, value=val)
    else:
        ws.append(list(headers))
        for row in rows:
            ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_multi_section_xlsx() -> bytes:
    """
    Return bytes of a .xlsx with two data sections separated by a blank row.

    Data rows use mixed types (string + number) so _detect_all_sections treats
    them as data rows rather than title rows, preserving the single-cell titles.

    Layout:
      Row 1:  "Q1 Results"          ← single-cell title
      Row 2:  "Jan", 100            ← data row (mixed types → not a title row)
      Row 3:  "Feb", 200
      Row 4:  "Mar", 150
      Row 5:  (empty)               ← section separator
      Row 6:  "Q2 Results"          ← single-cell title
      Row 7:  "Apr", 180
      Row 8:  "May", 220
      Row 9:  "Jun", 170
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Q1 Results"])                          # row 1: title
    for month, val in [("Jan", 100), ("Feb", 200), ("Mar", 150)]:
        ws.append([month, val])                        # rows 2-4: data
    ws.append([None])                                  # row 5: blank separator
    ws.append(["Q2 Results"])                          # row 6: title
    for month, val in [("Apr", 180), ("May", 220), ("Jun", 170)]:
        ws.append([month, val])                        # rows 7-9: data
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_percentage_xlsx() -> bytes:
    """
    Return bytes of a .xlsx with a decimal percentage column [0, 1].
    ChartService._detect_percentage_column should recognise this and
    build_plotly_data should auto-resolve to 'horizontal_bar'.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Percentage"])
    for cat, pct in [("A", 0.45), ("B", 0.78), ("C", 0.32)]:
        ws.append([cat, pct])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_scatter_xlsx() -> bytes:
    """Return bytes of a .xlsx with two numeric columns for scatter plots."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["X", "Y"])
    for x, y in [(1, 2), (3, 5), (6, 4), (9, 8)]:
        ws.append([x, y])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_pie_xlsx() -> bytes:
    """Return bytes of a .xlsx with label/value pairs for pie charts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Slice", "Value"])
    for label, val in [("Alpha", 30), ("Beta", 45), ("Gamma", 25)]:
        ws.append([label, val])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_ratings_xlsx() -> bytes:
    """Return bytes of a .xlsx with item/score pairs for ratings-mode testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item", "Score"])
    # 3.0 <= 3.15 → red, 3.5 in middle → yellow, 4.0 > 3.85 → green
    for item, score in [("X", 3.0), ("Y", 3.5), ("Z", 4.0)]:
        ws.append([item, score])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
